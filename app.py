import streamlit as st
import pandas as pd
import requests
import io
from SPARQLWrapper import SPARQLWrapper, JSON
from openpyxl.styles import PatternFill

# --- Fonctions de comparaison Excel ---
def load_excel_sheets(file) -> pd.ExcelFile:
    return pd.ExcelFile(file)

@st.cache_data
def read_titles(file_path: str, sheet_name: str) -> pd.Index:
    df = pd.read_excel(file_path, sheet_name=sheet_name, usecols=["Title"])
    return pd.Index(df["Title"])

@st.cache_data
def compare_titles(file_path: str, sheet_old: str, sheet_new: str, flag_col_idx: int):
    idx_old = read_titles(file_path, sheet_old)
    idx_new = read_titles(file_path, sheet_new)
    common   = idx_old.intersection(idx_new)
    excluded = idx_old.difference(idx_new)
    new      = idx_new.difference(idx_old)
    df_new = pd.read_excel(file_path, sheet_name=sheet_new)
    flag_series = df_new.iloc[:, flag_col_idx].astype(str)
    mask_yes = flag_series.str.upper().str.contains("YES", na=False)
    titles_flagged = df_new.loc[mask_yes, "Title"]
    common_yes = pd.Index(titles_flagged).intersection(common)
    return common, excluded, new, common_yes

@st.cache_data(ttl=24*3600)
def fetch_netflix_anime_titles_sparql() -> set[str]:
    endpoint = SPARQLWrapper("https://query.wikidata.org/sparql")
    endpoint.agent = "Streamlit/Wikidata Python"
    query = """
    SELECT ?itemLabel WHERE {
      ?item wdt:P31/wdt:P279* ?type;
            wdt:P750 wd:Q907311.
      VALUES ?type {
        wd:Q202866 wd:Q1107 wd:Q581714 wd:Q117467246 wd:Q17175676
        wd:Q4765080 wd:Q117467240 wd:Q113671041 wd:Q110955215 wd:Q54932319
      }
      SERVICE wikibase:label { bd:serviceParam wikibase:language "en". }
    }
    """
    endpoint.setQuery(query)
    endpoint.setReturnFormat(JSON)
    results = endpoint.query().convert()
    return {b["itemLabel"]["value"] for b in results["results"]["bindings"]}

# --- Streamlit UI ---
def main():
    st.title("Comparateur de Titres Netflix")
    st.markdown("Chargez un fichier Excel, comparez les titres et générez un rapport.")

    uploaded_file = st.file_uploader("Fichier Excel (xls/xlsx)", type=["xls", "xlsx"])
    if not uploaded_file:
        st.info("Importez un fichier pour démarrer.")
        return

    xls = load_excel_sheets(uploaded_file)
    col1, col2 = st.columns(2)
    with col1:
         sheet_old = st.selectbox("Ancienne feuille", xls.sheet_names)
    with col2:     
         sheet_new = st.selectbox(
                "Nouvelle feuille", xls.sheet_names, index=min(1, len(xls.sheet_names)-1)
    )

    col_number = st.number_input(
        "Numéro de la colonne du flag (1 pour A, 7 pour G, etc.)", 1, 50, 7
    )
    flag_col_idx = col_number - 1

    if st.button("Comparer"):
        common, excluded, new, common_yes = compare_titles(
            uploaded_file, sheet_old, sheet_new, flag_col_idx
        )
        st.session_state.update({
            'common': common,
            'new': new,
            'common_yes': common_yes
        })

    if 'common' in st.session_state:
        common = st.session_state['common']
        new = st.session_state['new']
        common_yes = st.session_state['common_yes']

        st.success("Résultats prêts !")
        st.write(f"Titres communs : {len(common)} ; Nouvelles entrées : {len(new)}")
        st.write(f"Parmi les titres communs, {len(common_yes)} sont flagués 'YES'.")

        # Préparer DataFrame global
        df_common = pd.DataFrame(common, columns=["Title"])
        df_common['Source'] = 'Common'
        df_new = pd.DataFrame(new, columns=["Title"])
        df_new['Source'] = 'New'
        df_all = pd.concat([df_common, df_new], ignore_index=True)
        df_all['Flagged'] = df_all['Title'].isin(common_yes)

        # Générer un fichier Excel stylé
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df_all.to_excel(writer, index=False, sheet_name='Report')
            ws = writer.sheets['Report']
            yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            for i, flagged in enumerate(df_all['Flagged'], start=2):
                if flagged:
                    for col in range(1, ws.max_column+1):
                        ws.cell(row=i, column=col).fill = yellow
        buffer.seek(0)

        st.download_button(
            label="Télécharger le rapport Excel",
            data=buffer,
            file_name='common_new_report.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

if __name__ == "__main__":
    main()
